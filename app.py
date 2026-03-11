import os
import calendar
import secrets
from datetime import date, datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from flask import (
    Flask, render_template, request, redirect, url_for, flash, jsonify, abort, send_file
)
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, Admin, Employee, WorkInterval, DailyAdjustment
from functools import wraps

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///worktime.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'admin_login'

HOLIDAYS = {
    (1, 1), (1, 2), (1, 3), (1, 4), (1, 5), (1, 6), (1, 7), (1, 8),
    (2, 23),
    (3, 8),
    (5, 1),
    (5, 9),
    (6, 12),
    (11, 4),
}

def get_total_minutes_for_day(employee_id, day_date):
    start = datetime.combine(day_date, datetime.min.time())
    end = start + timedelta(days=1)
    intervals = WorkInterval.query.filter(
        WorkInterval.employee_id == employee_id,
        WorkInterval.type == 'work',
        WorkInterval.start_time >= start,
        WorkInterval.start_time < end
    ).all()
    total_seconds = 0
    for interval in intervals:
        interval_end = interval.end_time or datetime.utcnow()
        total_seconds += (interval_end - interval.start_time).total_seconds()
    adjustments = DailyAdjustment.query.filter_by(employee_id=employee_id, date=day_date).all()
    total_minutes = total_seconds / 60 + sum(adj.delta_minutes for adj in adjustments)
    return int(total_minutes)

@app.context_processor
def utility_processor():
    return dict(date=date)

with app.app_context():
    db.create_all()
    if not Admin.query.filter_by(username='admin').first():
        admin = Admin(username='admin', password_hash=generate_password_hash('admin123'))
        db.session.add(admin)
        db.session.commit()

class AdminUser(UserMixin):
    def __init__(self, admin):
        self.id = admin.id

@login_manager.user_loader
def load_user(user_id):
    admin = Admin.query.get(int(user_id))
    return AdminUser(admin) if admin else None

def employee_required(f):
    @wraps(f)
    def decorated_function(token, *args, **kwargs):
        employee = Employee.query.filter_by(token=token).first()
        if not employee:
            abort(404)
        return f(employee=employee, *args, **kwargs)
    return decorated_function

# --- Маршруты для администратора ---
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        admin = Admin.query.filter_by(username=username).first()
        if admin and check_password_hash(admin.password_hash, password):
            login_user(AdminUser(admin))
            return redirect(url_for('admin_dashboard'))
        flash('Неверное имя пользователя или пароль')
    return render_template('admin/login.html')

@app.route('/admin/logout')
@login_required
def admin_logout():
    logout_user()
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    employees = Employee.query.all()
    today = date.today()
    start_of_month = date(today.year, today.month, 1)
    end_of_month = today

    employees_data = []
    for emp in employees:
        intervals = WorkInterval.query.filter(
            WorkInterval.employee_id == emp.id,
            WorkInterval.type == 'work',
            WorkInterval.start_time >= datetime.combine(start_of_month, datetime.min.time()),
            WorkInterval.start_time < datetime.combine(end_of_month + timedelta(days=1), datetime.min.time())
        ).all()
        total_seconds = 0
        for interval in intervals:
            end = interval.end_time or datetime.utcnow()
            total_seconds += (end - interval.start_time).total_seconds()
        total_minutes_from_intervals = total_seconds / 60

        adjustments = DailyAdjustment.query.filter(
            DailyAdjustment.employee_id == emp.id,
            DailyAdjustment.date >= start_of_month,
            DailyAdjustment.date <= end_of_month
        ).all()
        total_adjustment_minutes = sum(adj.delta_minutes for adj in adjustments)

        total_minutes = total_minutes_from_intervals + total_adjustment_minutes
        hours = int(total_minutes // 60)
        minutes = int(total_minutes % 60)

        # Форматирование телефона
        phone = emp.phone.strip()
        if len(phone) == 11 and phone.startswith('9'):
            phone_formatted = f"+7 ({phone[1:4]}) {phone[4:7]}-{phone[7:9]}-{phone[9:]}"
        elif len(phone) == 10:
            phone_formatted = f"+7 ({phone[0:3]}) {phone[3:6]}-{phone[6:8]}-{phone[8:]}"
        elif phone.startswith('7') and len(phone) == 11:
            phone_formatted = f"+7 ({phone[1:4]}) {phone[4:7]}-{phone[7:9]}-{phone[9:]}"
        elif phone.startswith('8') and len(phone) == 11:
            phone_formatted = f"+7 ({phone[1:4]}) {phone[4:7]}-{phone[7:9]}-{phone[9:]}"
        else:
            phone_formatted = phone

        employees_data.append({
            'emp': emp,
            'total_time': f"{hours}:{minutes:02d}",
            'phone_formatted': phone_formatted
        })

    return render_template('admin/dashboard.html', employees_data=employees_data)

@app.route('/admin/employees')
@login_required
def admin_employees():
    employees = Employee.query.all()
    return render_template('admin/employees.html', employees=employees)

@app.route('/admin/employees/add', methods=['POST'])
@login_required
def add_employee():
    first_name = request.form['first_name']
    last_name = request.form['last_name']
    middle_name = request.form.get('middle_name', '')
    phone = request.form['phone']
    employee = Employee(first_name=first_name, last_name=last_name,
                        middle_name=middle_name, phone=phone)
    db.session.add(employee)
    db.session.commit()
    link = url_for('employee_panel', token=employee.token, _external=True)
    flash(f'Сотрудник добавлен. Ссылка для него: {link}')
    return redirect(url_for('admin_employees'))

@app.route('/admin/reports', methods=['GET', 'POST'])
@login_required
def admin_reports():
    employees = Employee.query.all()
    report_data = []
    if request.method == 'POST':
        start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d')
        end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d') + timedelta(days=1)
        for emp in employees:
            intervals = WorkInterval.query.filter(
                WorkInterval.employee_id == emp.id,
                WorkInterval.type == 'work',
                WorkInterval.start_time >= start_date,
                WorkInterval.start_time < end_date
            ).all()
            total_seconds = 0
            for interval in intervals:
                end = interval.end_time or datetime.utcnow()
                if end > start_date:
                    end = min(end, end_date)
                if interval.start_time < end_date:
                    start = max(interval.start_time, start_date)
                    total_seconds += (end - start).total_seconds()
            total_minutes_from_intervals = total_seconds / 60
            adjustments = DailyAdjustment.query.filter(
                DailyAdjustment.employee_id == emp.id,
                DailyAdjustment.date >= start_date.date(),
                DailyAdjustment.date < end_date.date()
            ).all()
            total_adjustment_minutes = sum(adj.delta_minutes for adj in adjustments)
            total_minutes = total_minutes_from_intervals + total_adjustment_minutes
            hours = int(total_minutes // 60)
            minutes = int(total_minutes % 60)
            report_data.append({
                'employee': f"{emp.last_name} {emp.first_name} {emp.middle_name}",
                'total_str': f"{hours}:{minutes:02d}"
            })
    return render_template('admin/reports.html', employees=employees, report_data=report_data)

MONTH_NAMES = {
    1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
    5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
    9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
}

@app.route('/admin/employee/<int:employee_id>/calendar')
@login_required
def employee_calendar(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    try:
        year = int(request.args.get('year', date.today().year))
        month = int(request.args.get('month', date.today().month))
        if month < 1 or month > 12:
            raise ValueError
    except:
        year = date.today().year
        month = date.today().month

    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year+1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month+1, 1) - timedelta(days=1)

    start_dt = datetime.combine(first_day, datetime.min.time())
    end_dt = datetime.combine(last_day + timedelta(days=1), datetime.min.time())

    intervals = WorkInterval.query.filter(
        WorkInterval.employee_id == employee_id,
        WorkInterval.type == 'work',
        WorkInterval.start_time >= start_dt,
        WorkInterval.start_time < end_dt
    ).all()

    adjustments = DailyAdjustment.query.filter(
        DailyAdjustment.employee_id == employee_id,
        DailyAdjustment.date >= first_day,
        DailyAdjustment.date <= last_day
    ).all()

    daily_minutes = {}
    daily_delta = {}
    daily_comments = {}

    for interval in intervals:
        day = interval.start_time.date()
        end = interval.end_time or datetime.utcnow()
        minutes = (end - interval.start_time).total_seconds() / 60
        daily_minutes[day] = daily_minutes.get(day, 0) + minutes

    for adj in adjustments:
        day = adj.date
        daily_minutes[day] = daily_minutes.get(day, 0) + adj.delta_minutes
        daily_delta[day] = daily_delta.get(day, 0) + adj.delta_minutes
        if adj.comment:
            daily_comments.setdefault(day, []).append(adj.comment)

    daily_minutes = {day: int(minutes) for day, minutes in daily_minutes.items()}

    cal = calendar.monthcalendar(year, month)

    month_data = {
        'year': year,
        'month': month,
        'cal': cal,
        'daily_minutes': daily_minutes,
        'daily_delta': daily_delta,
        'daily_comments': daily_comments
    }

    prev_year, prev_month = (year-1, 12) if month == 1 else (year, month-1)
    next_year, next_month = (year+1, 1) if month == 12 else (year, month+1)

    return render_template('admin/calendar.html',
                           employee=employee,
                           month_data=month_data,
                           prev_year=prev_year, prev_month=prev_month,
                           next_year=next_year, next_month=next_month,
                           holidays=HOLIDAYS,
                           month_name=MONTH_NAMES[month])

@app.route('/admin/employee/<int:employee_id>/adjustments')
@login_required
def get_adjustments(employee_id):
    date_str = request.args.get('date')
    if not date_str:
        return jsonify([])
    try:
        day_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except:
        return jsonify([])
    adjustments = DailyAdjustment.query.filter_by(employee_id=employee_id, date=day_date).all()
    return jsonify([{
        'id': adj.id,
        'delta_minutes': adj.delta_minutes,
        'comment': adj.comment,
        'created_at': adj.created_at.strftime('%Y-%m-%d %H:%M')
    } for adj in adjustments])

@app.route('/admin/employee/<int:employee_id>/adjust', methods=['POST'])
@login_required
def add_adjustment(employee_id):
    data = request.get_json()
    date_str = data.get('date')
    hours = int(data.get('hours', 0))
    minutes = int(data.get('minutes', 0))
    comment = data.get('comment', '')

    delta = hours * 60 + minutes
    if delta == 0 and not comment:
        return jsonify({'error': 'Нет изменений'}), 400

    day_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    adj = DailyAdjustment(employee_id=employee_id, date=day_date,
                          delta_minutes=delta, comment=comment)
    db.session.add(adj)
    db.session.commit()

    total = get_total_minutes_for_day(employee_id, day_date)
    hours_total = total // 60
    minutes_total = total % 60
    return jsonify({'success': True, 'total': f'{hours_total}:{minutes_total:02d}'})

# --- Маршруты для сотрудников ---
@app.route('/employee/<token>')
@employee_required
def employee_panel(employee):
    active_interval = WorkInterval.query.filter_by(employee_id=employee.id, end_time=None).first()
    status = 'none'
    if active_interval:
        status = active_interval.type
    return render_template('employee/panel.html', employee=employee, status=status)

@app.route('/api/employee/<token>/start', methods=['POST'])
@employee_required
def api_start(employee):
    active = WorkInterval.query.filter_by(employee_id=employee.id, end_time=None).first()
    if active:
        return jsonify({'error': 'Уже есть активная сессия'}), 400
    interval = WorkInterval(employee_id=employee.id, type='work')
    db.session.add(interval)
    db.session.commit()
    return jsonify({'status': 'work'})

@app.route('/api/employee/<token>/pause', methods=['POST'])
@employee_required
def api_pause(employee):
    active = WorkInterval.query.filter_by(employee_id=employee.id, end_time=None).first()
    if not active or active.type != 'work':
        return jsonify({'error': 'Нет активной работы для паузы'}), 400
    active.end_time = datetime.utcnow()
    pause = WorkInterval(employee_id=employee.id, type='pause')
    db.session.add(pause)
    db.session.commit()
    return jsonify({'status': 'pause'})

@app.route('/api/employee/<token>/resume', methods=['POST'])
@employee_required
def api_resume(employee):
    active = WorkInterval.query.filter_by(employee_id=employee.id, end_time=None).first()
    if not active or active.type != 'pause':
        return jsonify({'error': 'Нет активной паузы для возобновления'}), 400
    active.end_time = datetime.utcnow()
    work = WorkInterval(employee_id=employee.id, type='work')
    db.session.add(work)
    db.session.commit()
    return jsonify({'status': 'work'})

@app.route('/api/employee/<token>/stop', methods=['POST'])
@employee_required
def api_stop(employee):
    active = WorkInterval.query.filter_by(employee_id=employee.id, end_time=None).first()
    if not active:
        return jsonify({'error': 'Нет активной сессии'}), 400
    active.end_time = datetime.utcnow()
    db.session.commit()
    return jsonify({'status': 'none'})

# --- Маршруты для редактирования и удаления сотрудников ---
@app.route('/admin/employee/<int:employee_id>/data')
@login_required
def get_employee_data(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    return jsonify({
        'id': employee.id,
        'last_name': employee.last_name,
        'first_name': employee.first_name,
        'middle_name': employee.middle_name,
        'phone': employee.phone
    })

@app.route('/admin/employee/<int:employee_id>/edit', methods=['POST'])
@login_required
def edit_employee(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    data = request.get_json()
    employee.last_name = data['last_name']
    employee.first_name = data['first_name']
    employee.middle_name = data.get('middle_name', '')
    employee.phone = data['phone']
    if data.get('generate_new_token'):
        employee.token = secrets.token_urlsafe(32)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/employee/<int:employee_id>/delete', methods=['POST'])
@login_required
def delete_employee(employee_id):
    try:
        employee = Employee.query.get_or_404(employee_id)
        db.session.delete(employee)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# --- Маршрут для экспорта отчёта в Excel ---
@app.route('/admin/reports/export', methods=['POST'])
@login_required
def export_reports_excel():
    start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d')
    end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d') + timedelta(days=1)
    employees = Employee.query.all()

    # Список дат в периоде
    date_list = []
    current = start_date.date()
    while current < end_date.date():
        date_list.append(current)
        current += timedelta(days=1)

    # Формируем название периода (для имени файла)
    month_names = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
                   'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
    last_day_of_month = (start_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    if start_date.day == 1 and (end_date - timedelta(days=1)).day == last_day_of_month.day:
        month_name = month_names[start_date.month - 1]
        year = start_date.year
        filename = f"часы работы за {month_name} {year}.xlsx"
    else:
        filename = f"часы работы с {start_date.strftime('%d.%m.%Y')} по {(end_date - timedelta(days=1)).strftime('%d.%m.%Y')}.xlsx"

    # Собираем данные по каждому сотруднику
    data = []  # список кортежей (employee, time_in_list, time_out_list, total_seconds_list)
    for emp in employees:
        time_in = []   # время прихода для каждого дня (datetime.time или None)
        time_out = []  # время ухода (datetime.time или None)
        total_seconds_day = []  # общее количество секунд за день (только по завершённым интервалам)
        for day in date_list:
            # Находим все рабочие интервалы за день
            day_start = datetime.combine(day, datetime.min.time())
            day_end = day_start + timedelta(days=1)
            intervals = WorkInterval.query.filter(
                WorkInterval.employee_id == emp.id,
                WorkInterval.type == 'work',
                WorkInterval.start_time >= day_start,
                WorkInterval.start_time < day_end
            ).order_by(WorkInterval.start_time).all()

            if intervals:
                # Приход = начало первого интервала
                first = intervals[0].start_time
                time_in.append(first.time())
                # Уход = конец последнего интервала (только если завершён)
                last = intervals[-1]
                if last.end_time:
                    time_out.append(last.end_time.time())
                else:
                    time_out.append(None)
                # Общее время за день (сумма завершённых интервалов)
                total_sec = 0
                for inv in intervals:
                    if inv.end_time:
                        total_sec += (inv.end_time - inv.start_time).total_seconds()
                total_seconds_day.append(total_sec)
            else:
                time_in.append(None)
                time_out.append(None)
                total_seconds_day.append(0)
        data.append((emp, time_in, time_out, total_seconds_day))

    # Создаём книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт по дням"

    # Стили
    center_alignment = Alignment(horizontal='center', vertical='center')
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    bold_font = Font(bold=True)
    time_format = 'HH:MM:SS'  # формат времени

    # Определяем колонки:
    # A - ФИО (только в строке Приход)
    # B - легенда (Приход/Уход/Часы)
    # C ... - даты
    fio_col = 1          # A
    legend_col = 2       # B
    first_data_col = 3   # C
    last_data_col = first_data_col + len(date_list) - 1
    total_col = last_data_col + 1

    # Заголовки (строка 1)
    ws.cell(row=1, column=fio_col, value="Ф И О")
    ws.cell(row=1, column=legend_col, value="")  # пусто в легенде
    for col, day in enumerate(date_list, start=first_data_col):
        # дата в формате ГГГГ-ММ-ДД 00:00:00
        ws.cell(row=1, column=col, value=day.strftime('%Y-%m-%d 00:00:00'))
    ws.cell(row=1, column=total_col, value="часы сумма")

    # Дни недели (строка 2)
    weekdays = ['вс', 'пн', 'вт', 'ср', 'чт', 'пт', 'сб']
    for col, day in enumerate(date_list, start=first_data_col):
        cell = ws.cell(row=2, column=col, value=weekdays[day.weekday()])
        cell.alignment = center_alignment
        if day.weekday() >= 5:  # выходные
            cell.fill = gray_fill

    # Данные сотрудников
    current_row = 3  # начнём с 3-й строки
    for emp, time_in, time_out, total_seconds_day in data:
        fio = f"{emp.last_name} {emp.first_name} {emp.middle_name or ''}".strip()

        # Строка "Приход"
        ws.cell(row=current_row, column=fio_col, value=fio).alignment = center_alignment
        ws.cell(row=current_row, column=legend_col, value="Приход").alignment = center_alignment
        for col, ti in enumerate(time_in, start=first_data_col):
            if ti:
                cell = ws.cell(row=current_row, column=col, value=ti.strftime('%H:%M:%S'))
                cell.number_format = time_format
                cell.alignment = center_alignment
        current_row += 1

        # Строка "Уход"
        ws.cell(row=current_row, column=legend_col, value="Уход").alignment = center_alignment
        for col, to in enumerate(time_out, start=first_data_col):
            if to:
                cell = ws.cell(row=current_row, column=col, value=to.strftime('%H:%M:%S'))
                cell.number_format = time_format
                cell.alignment = center_alignment
        current_row += 1

        # Строка "Часы"
        ws.cell(row=current_row, column=legend_col, value="Часы").alignment = center_alignment
        for col, sec in enumerate(total_seconds_day, start=first_data_col):
            if sec > 0:
                hours = int(sec // 3600)
                minutes = int((sec % 3600) // 60)
                seconds = int(sec % 60)
                cell = ws.cell(row=current_row, column=col, value=f"{hours:02d}:{minutes:02d}:{seconds:02d}")
                cell.alignment = center_alignment
        # Итоговая сумма часов для этого сотрудника (в колонке total_col)
        total_sec_period = sum(total_seconds_day)
        if total_sec_period > 0:
            hours = int(total_sec_period // 3600)
            minutes = int((total_sec_period % 3600) // 60)
            seconds = int(total_sec_period % 60)
            cell = ws.cell(row=current_row, column=total_col, value=f"{hours:02d}:{minutes:02d}:{seconds:02d}")
            cell.alignment = center_alignment
            cell.font = bold_font
        current_row += 1

    # Настройка ширины колонок
    ws.column_dimensions[get_column_letter(fio_col)].width = 25    # ФИО
    ws.column_dimensions[get_column_letter(legend_col)].width = 8  # легенда
    for col in range(first_data_col, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True)